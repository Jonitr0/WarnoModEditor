import openpyxl
from string import digits


class SheetManager:
    def __init__(self, xlsx_path: str, sheet_name: str, prefix: str):
        wb_obj = openpyxl.open(xlsx_path)
        self.sheet = wb_obj.get_sheet_by_name(sheet_name)

        self.object_name_to_index = {}

        # save row index for each object to be edited
        for i in range(2, self.sheet.max_row + 1):
            self.object_name_to_index[prefix + str(self.sheet.cell(row=i, column=1).value)] = i

        self.category_to_index = {}

        # save column index for each category
        for i in range(2, self.sheet.max_column + 1):
            self.category_to_index[str(self.sheet.cell(row=1, column=i).value)] = i

    def get_attribute_for_object(self, obj_name: str, attr_name: str):
        row = self.object_name_to_index[obj_name]
        column = self.category_to_index[attr_name]

        return self.sheet.cell(row=row, column=column).value

    def contains_object(self, obj_name: str):
        return self.object_name_to_index.__contains__(obj_name)


class SheetMultiManager:
    def __init__(self, xlsx_path: str, sheet_name: str, prefix: str):
        wb_obj = openpyxl.open(xlsx_path)
        self.sheet = wb_obj.get_sheet_by_name(sheet_name)

        self.object_name_to_index = {}

        # save row index for each object to be edited
        for i in range(2, self.sheet.max_row + 1):
            name = prefix + str(self.sheet.cell(row=i, column=1).value)
            if not self.object_name_to_index.__contains__(name):
                obj_index = 0
                self.object_name_to_index[name] = {}
            else:
                obj_index = len(self.object_name_to_index[name])
            self.object_name_to_index[name][obj_index] = i

        self.category_to_index = {}

        # save column index for each category
        for i in range(2, self.sheet.max_column + 1):
            self.category_to_index[str(self.sheet.cell(row=1, column=i).value)] = i

    def get_attribute_for_object(self, obj_name: str, attr_name: str, obj_index: int):
        row = self.object_name_to_index[obj_name][obj_index]
        column = self.category_to_index[attr_name]

        return self.sheet.cell(row=row, column=column).value

    def number_entries_for_object(self, obj_name: str):
        if not self.object_name_to_index.__contains__(obj_name):
            return 0
        return len(self.object_name_to_index[obj_name])


def edit_units_from_xlsx(unit_desc, xlsx_path: str = "Mod.xlsx"):
    sheet_manager = SheetManager(xlsx_path, sheet_name="Units", prefix="Descriptor_Unit_")

    for unit in unit_desc:
        name = unit.namespace
        unit_obj = unit.value

        if not sheet_manager.contains_object(name):
            continue

        for module in unit_obj.by_member("ModulesDescriptors").value:
            try:
                module_type = module.value.type
            except Exception:
                continue
            if module_type == "TModuleSelector":
                try:
                    default = module.value.by_member("Default").value.type
                except Exception:
                    continue
                match default:
                    # set HP
                    case "TBaseDamageModuleDescriptor":
                        hp = sheet_manager.get_attribute_for_object(name, "hp")
                        if not hp:
                            continue
                        hp_module = module.value.by_member("Default").value.by_member("MaxPhysicalDamages")
                        hp_module.value = int(hp)
                    # set armor
                    case "TDamageModuleDescriptor":
                        armor_module = module.value.by_member("Default").value.by_member("BlindageProperties").value
                        for aspect in ["Front", "Sides", "Rear", "Top"]:
                            armor_value = sheet_manager.get_attribute_for_object(name, aspect + "_armor_value")
                            if not armor_value:
                                continue
                            armor_type = sheet_manager.get_attribute_for_object(name, aspect + "_armor_type")
                            if not armor_type:
                                armor_type = armor_module.by_member("ArmorDescriptor" + aspect).value.rstrip(digits)
                            armor_module.by_member("ArmorDescriptor" + aspect).value = armor_type + str(armor_value)
                    # set price
                    case "TProductionModuleDescriptor":
                        price = sheet_manager.get_attribute_for_object(name, "price")
                        if not price:
                            continue
                        price_module = module.value.by_member("Default").value.by_member(
                            "ProductionRessourcesNeeded").value
                        price_module.by_key("~/Resource_CommandPoints").value = int(price)
                    # optics
                    case "TScannerConfigurationDescriptor":
                        optics = sheet_manager.get_attribute_for_object(name, "optics")
                        if not optics:
                            continue
                        scanner = module.value.by_member("Default").value.by_member("OpticalStrength")
                        scanner_alt = module.value.by_member("Default").value.by_member("OpticalStrengthAltitude")
                        match optics:
                            case "bad":
                                scanner.value = 42.45
                                scanner_alt.value = 10
                            case "mediocre":
                                scanner.value = 63.675
                                scanner_alt.value = 20
                            case "normal":
                                scanner.value = 84.9
                                scanner_alt.value = 20
                            case _:
                                pass
                    case _:
                        pass
            elif module_type == "TReverseScannerWithIdentificationDescriptor":
                optics = sheet_manager.get_attribute_for_object(name, "optics")
                if not optics:
                    continue

                vis_module = module.value.by_member("VisibilityRollRule").value
                id_chance = vis_module.by_member("IdentifyBaseProbability")
                id_rate = vis_module.by_member("TimeBetweenEachIdentifyRoll")

                match optics:
                    case "bad":
                        id_chance.value = 0.13
                        id_rate.value = 15.0
                    case "mediocre":
                        id_chance.value = 0.2
                        id_rate.value = 10.0
                    case "normal":
                        id_chance.value = 0.26
                        id_rate.value = 8.0
                    case _:
                        pass


# set ammo for given unit on given turret
def edit_turrets_from_xlsx(weapon_desc, xlsx_path: str = "Mod.xlsx"):
    sheet_manager = SheetMultiManager(xlsx_path, sheet_name="Turrets", prefix="WeaponDescriptor_")

    for weapon in weapon_desc:
        name = weapon.namespace
        weapon_obj = weapon.value

        for i in range(sheet_manager.number_entries_for_object(name)):
            try:
                turret_index = int(sheet_manager.get_attribute_for_object(name, "turret_index", i))
                weapon_index = int(sheet_manager.get_attribute_for_object(name, "weapon_index", i))
            except TypeError:
                continue
            ammo = sheet_manager.get_attribute_for_object(name, "ammo_name", i)
            if not ammo:
                continue

            turret = weapon_obj.by_member("TurretDescriptorList").value[turret_index]
            mount = turret.value.by_member("MountedWeaponDescriptorList").value[weapon_index]
            mount.value.by_member("Ammunition").value = "~/Ammo_" + ammo


# TODO: read this from BaseGDConstantes
RANGE_FACTOR = 2.83


def edit_ammo_from_xlsx(ammo_desc, xlsx_path: str = "Mod.xlsx"):
    sheet_manager = SheetManager(xlsx_path, sheet_name="Ammo", prefix="Ammo_")

    for ammo in ammo_desc:
        name = ammo.namespace

        if not sheet_manager.contains_object(name):
            continue

        ammo_obj = ammo.value

        arme_value = sheet_manager.get_attribute_for_object(name, "arme_value")
        if arme_value:
            ammo_obj.by_member("Arme").value.by_member("Index").value = arme_value
        arme_family = sheet_manager.get_attribute_for_object(name, "arme_family")
        if arme_family:
            ammo_obj.by_member("Arme").value.by_member("Family").value = arme_family

        range_ground = sheet_manager.get_attribute_for_object(name, "range_ground")
        if range_ground:
            range_adjusted = round(float(range_ground) * RANGE_FACTOR)
            ammo_obj.by_member("PorteeMaximale").value = "(({range}) * Metre)".format(range=range_adjusted)

        acc_stat = sheet_manager.get_attribute_for_object(name, "acc_stationary")
        acc_mov = sheet_manager.get_attribute_for_object(name, "acc_moving")

        hit_chance = ammo_obj.by_member("HitRollRuleDescriptor").value.by_member("BaseHitValueModifiers").value
        for entry in hit_chance:
            if acc_stat and entry.value[0] == "EBaseHitValueModifier/Idling":
                entry.value = ("EBaseHitValueModifier/Idling", acc_stat)
            elif acc_mov and entry.value[0] == "EBaseHitValueModifier/Moving":
                entry.value = ("EBaseHitValueModifier/Moving", acc_mov)


